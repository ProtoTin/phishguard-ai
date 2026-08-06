"""Normalize, deduplicate, split, and report on the declared datasets."""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import unicodedata
import zipfile
from collections import Counter, defaultdict
from collections.abc import Iterable, Iterator
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Literal, cast
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import load_workbook

from phishguard.data.download import verify_file
from phishguard.data.manifest import DatasetSource, Label, load_sources

Split = Literal["train", "validation", "test", "external_test"]


@dataclass(frozen=True)
class RawRecord:
    """A source record after label mapping and basic input cleanup."""

    content_type: Literal["email", "url"]
    text: str
    label: Label
    source: str
    source_record_id: str
    group: str
    split_policy: str


@dataclass(frozen=True)
class ProcessedRecord:
    """Stable public schema written to processed JSON Lines files."""

    id: str
    content_type: Literal["email", "url"]
    text: str
    label: Label
    source: str
    source_record_id: str
    group_id: str
    split: Split


@dataclass
class SourceStats:
    """Aggregate preparation counts for one source."""

    raw_rows: int = 0
    accepted_rows: int = 0
    retained_rows: int = 0
    excluded_labels: int = 0
    empty_rows: int = 0


def clean_text(value: object) -> str:
    """Normalize line endings and control characters while preserving content."""

    text = unicodedata.normalize("NFC", str(value or ""))
    return text.replace("\x00", "").replace("\r\n", "\n").replace("\r", "\n").strip()


def canonical_email(text: str) -> str:
    """Build a conservative near-exact key insensitive to case and punctuation."""

    normalized = unicodedata.normalize("NFKC", text).casefold()
    return " ".join(re.findall(r"[\w@]+", normalized, flags=re.UNICODE))


def canonical_url(value: str) -> str:
    """Build an offline URL key without resolving or requesting the destination."""

    candidate = clean_text(value)
    parsed = urlsplit(candidate if "://" in candidate else f"http://{candidate}")
    hostname = (parsed.hostname or "").casefold().rstrip(".")
    if hostname.startswith("www."):
        hostname = hostname[4:]
    port = parsed.port
    if port in {80, 443}:
        port = None
    authority = f"{hostname}:{port}" if port else hostname
    path = re.sub(r"/{2,}", "/", parsed.path or "/").rstrip("/") or "/"
    query = urlencode(sorted(parse_qsl(parsed.query, keep_blank_values=True)))
    return urlunsplit(("", authority, path, query, ""))


def fingerprint(record: RawRecord) -> str:
    canonical = (
        canonical_email(record.text)
        if record.content_type == "email"
        else canonical_url(record.text)
    )
    return hashlib.sha256(f"{record.content_type}\0{canonical}".encode()).hexdigest()


def stable_bucket(seed: int, content_type: str, group: str) -> int:
    """Map a group deterministically into one of 10,000 buckets."""

    payload = f"{seed}\0{content_type}\0{group}".encode()
    return int.from_bytes(hashlib.sha256(payload).digest()[:8], "big") % 10_000


def assign_split(record: RawRecord, seed: int) -> Split:
    """Assign a grouped deterministic split according to source policy."""

    if record.split_policy == "external_test":
        return "external_test"
    bucket = stable_bucket(seed, record.content_type, record.group)
    if record.split_policy == "train_validation":
        return "train" if bucket < 8_000 else "validation"
    if bucket < 7_000:
        return "train"
    if bucket < 8_500:
        return "validation"
    return "test"


def _mapped_record(
    source: DatasetSource,
    row_number: int,
    text: object,
    source_label: object,
    group: object,
    stats: SourceStats,
) -> RawRecord | None:
    stats.raw_rows += 1
    cleaned = clean_text(text)
    if not cleaned:
        stats.empty_rows += 1
        return None
    label = source.label_mapping.get(clean_text(source_label))
    if label is None:
        stats.excluded_labels += 1
        return None
    stats.accepted_rows += 1
    return RawRecord(
        content_type=source.content_type,
        text=cleaned,
        label=label,
        source=source.id,
        source_record_id=str(row_number),
        group=f"{source.id}:{clean_text(group) or row_number}",
        split_policy=source.split_policy,
    )


def load_social_engineering_xlsx(
    source: DatasetSource, path: Path, stats: SourceStats
) -> Iterator[RawRecord]:
    """Read the Zenodo workbook whose text and label share a tab-delimited cell."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        if worksheet is None:
            raise ValueError(f"Workbook {path.name} contains no active worksheet")
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_number == 1:
                continue
            combined = "\t".join(clean_text(value) for value in row if value is not None)
            if "\t" not in combined:
                stats.raw_rows += 1
                stats.empty_rows += 1
                continue
            text, source_label = combined.rsplit("\t", 1)
            record = _mapped_record(source, row_number, text, source_label, str(row_number), stats)
            if record:
                yield record
    finally:
        workbook.close()


def load_validation_csv(
    source: DatasetSource, path: Path, stats: SourceStats
) -> Iterator[RawRecord]:
    """Read the dedicated external email validation dataset."""

    with path.open(encoding="utf-8-sig", newline="") as source_file:
        reader = csv.DictReader(source_file)
        required = {"Email Text", "Email Type"}
        if not reader.fieldnames or not required.issubset(reader.fieldnames):
            raise ValueError(f"Unexpected columns in {path.name}")
        for row_number, row in enumerate(reader, start=2):
            record = _mapped_record(
                source,
                row_number,
                row["Email Text"],
                row["Email Type"],
                str(row_number),
                stats,
            )
            if record:
                yield record


def load_curated_email_csv(
    source: DatasetSource, path: Path, stats: SourceStats
) -> Iterator[RawRecord]:
    """Read subject, body, and binary labels from the curated Enron CSV."""

    csv.field_size_limit(10 * 1024 * 1024)
    with path.open(encoding="utf-8-sig", errors="replace", newline="") as source_file:
        reader = csv.DictReader(source_file)
        required = {"subject", "body", "label"}
        if not reader.fieldnames or not required.issubset(reader.fieldnames):
            raise ValueError(f"Unexpected columns in {path.name}")
        for row_number, row in enumerate(reader, start=2):
            subject = clean_text(row["subject"])
            body = clean_text(row["body"])
            text = f"Subject: {subject}\n\n{body}" if subject else body
            group = canonical_email(subject) if subject else str(row_number)
            record = _mapped_record(source, row_number, text, row["label"], group, stats)
            if record:
                yield record


def load_phiusiil_zip(source: DatasetSource, path: Path, stats: SourceStats) -> Iterator[RawRecord]:
    """Read raw URL text and labels from the UCI archive without visiting URLs."""

    member = "PhiUSIIL_Phishing_URL_Dataset.csv"
    with zipfile.ZipFile(path) as archive:
        if member not in archive.namelist():
            raise ValueError(f"Expected {member} in {path.name}")
        with archive.open(member) as binary_file:
            text_file = io.TextIOWrapper(binary_file, encoding="utf-8-sig", newline="")
            reader = csv.DictReader(text_file)
            required = {"URL", "Domain", "label"}
            if not reader.fieldnames or not required.issubset(reader.fieldnames):
                raise ValueError(f"Unexpected columns in {member}")
            for row_number, row in enumerate(reader, start=2):
                record = _mapped_record(
                    source,
                    row_number,
                    row["URL"],
                    row["label"],
                    canonical_url(row["Domain"]),
                    stats,
                )
                if record:
                    yield record


def load_records(source: DatasetSource, path: Path, stats: SourceStats) -> Iterator[RawRecord]:
    """Dispatch a pinned source to its schema-specific reader."""

    if source.format == "curated_email_csv":
        yield from load_curated_email_csv(source, path, stats)
    elif source.format == "zenodo_tsv_in_xlsx":
        yield from load_social_engineering_xlsx(source, path, stats)
    elif source.format == "zenodo_validation_csv":
        yield from load_validation_csv(source, path, stats)
    elif source.format == "uci_phiusiil_zip_csv":
        yield from load_phiusiil_zip(source, path, stats)
    else:
        raise ValueError(f"Unsupported dataset format: {source.format}")


def deduplicate(records: Iterable[RawRecord]) -> tuple[list[RawRecord], int, int]:
    """Remove canonical duplicates and every example involved in a label conflict."""

    by_fingerprint: dict[str, RawRecord] = {}
    conflicts: set[str] = set()
    duplicates = 0
    for record in records:
        key = fingerprint(record)
        existing = by_fingerprint.get(key)
        if existing is None:
            by_fingerprint[key] = record
        elif existing.label == record.label:
            duplicates += 1
        else:
            conflicts.add(key)
    for key in conflicts:
        by_fingerprint.pop(key, None)
    return list(by_fingerprint.values()), duplicates, len(conflicts)


def process_records(records: Iterable[RawRecord], seed: int) -> list[ProcessedRecord]:
    """Create stable identifiers and grouped splits for deduplicated records."""

    processed = []
    for record in records:
        record_fingerprint = fingerprint(record)
        split = assign_split(record, seed)
        group_id = hashlib.sha256(f"{record.content_type}\0{record.group}".encode()).hexdigest()[
            :16
        ]
        processed.append(
            ProcessedRecord(
                id=record_fingerprint[:20],
                content_type=record.content_type,
                text=record.text,
                label=record.label,
                source=record.source,
                source_record_id=record.source_record_id,
                group_id=group_id,
                split=split,
            )
        )
    return sorted(processed, key=lambda item: (item.split, item.content_type, item.id))


def write_outputs(records: list[ProcessedRecord], output_dir: Path, seed: int) -> dict[str, object]:
    """Write normalized split files and return their aggregate manifest."""

    output_dir.mkdir(parents=True, exist_ok=True)
    by_split: dict[Split, list[ProcessedRecord]] = defaultdict(list)
    for record in records:
        by_split[record.split].append(record)

    split_counts: dict[str, dict[str, int]] = {}
    for split in ("train", "validation", "test", "external_test"):
        split_records = by_split[split]
        destination = output_dir / f"{split}.jsonl"
        with destination.open("w", encoding="utf-8", newline="\n") as output:
            for record in split_records:
                output.write(json.dumps(asdict(record), ensure_ascii=False) + "\n")
        counts = Counter(f"{item.content_type}:{item.label}" for item in split_records)
        split_counts[split] = dict(sorted(counts.items()))

    manifest: dict[str, object] = {
        "schema_version": 1,
        "seed": seed,
        "total_records": len(records),
        "splits": split_counts,
        "fields": [
            "id",
            "content_type",
            "text",
            "label",
            "source",
            "source_record_id",
            "group_id",
            "split",
        ],
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, indent=2) + "\n", encoding="utf-8"
    )
    return manifest


def build_quality_report(
    sources: list[DatasetSource],
    source_stats: dict[str, SourceStats],
    manifest: dict[str, object],
    duplicates: int,
    label_conflicts: int,
) -> dict[str, object]:
    """Build an aggregate report containing no message or URL samples."""

    return {
        "report_version": 1,
        "sources": {
            source.id: {
                **asdict(source_stats[source.id]),
                "title": source.title,
                "content_type": source.content_type,
                "role": source.role,
                "license": source.license,
                "landing_page": source.landing_page,
            }
            for source in sources
        },
        "deduplication": {
            "same_label_duplicates_removed": duplicates,
            "conflicting_fingerprints_removed": label_conflicts,
        },
        "processed": manifest,
    }


def write_reports(report: dict[str, object], json_path: Path, markdown_path: Path) -> None:
    """Write machine-readable and reviewer-friendly aggregate reports."""

    json_path.parent.mkdir(parents=True, exist_ok=True)
    markdown_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")

    sources = require_dict(report["sources"], "sources")
    processed = require_dict(report["processed"], "processed")
    dedupe = require_dict(report["deduplication"], "deduplication")
    lines = [
        "# Data Quality Report",
        "",
        "> Generated by `python scripts/prepare_data.py`. It contains aggregate counts only.",
        "",
        "## Source processing",
        "",
        "| Source | Raw | Accepted | Retained | Excluded labels | Empty |",
        "| --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for source_id, value in sources.items():
        value = require_dict(value, f"source {source_id}")
        lines.append(
            f"| `{source_id}` | {value['raw_rows']} | {value['accepted_rows']} | "
            f"{value['retained_rows']} | {value['excluded_labels']} | "
            f"{value['empty_rows']} |"
        )
    lines.extend(
        [
            "",
            "## Deduplication",
            "",
            f"- Same-label canonical duplicates removed: {dedupe['same_label_duplicates_removed']}",
            "- Conflicting canonical fingerprints removed: "
            f"{dedupe['conflicting_fingerprints_removed']}",
            "",
            "Canonical matching is insensitive to Unicode normalization, case, whitespace, and",
            "punctuation for email; URL matching also normalizes scheme, host, default ports,",
            "path separators, fragments, and query ordering. Domain groups stay in one split.",
            "",
            "## Processed splits",
            "",
            "| Split | Email legitimate | Email phishing | URL legitimate | URL phishing |",
            "| --- | ---: | ---: | ---: | ---: |",
        ]
    )
    splits = require_dict(processed["splits"], "splits")
    for split, counts in splits.items():
        counts = require_dict(counts, f"split {split}")
        lines.append(
            f"| {split} | {counts.get('email:legitimate', 0)} | "
            f"{counts.get('email:phishing', 0)} | "
            f"{counts.get('url:legitimate', 0)} | {counts.get('url:phishing', 0)} |"
        )
    lines.extend(
        [
            "",
            f"Total processed records: **{processed['total_records']}**.",
            "",
            "## Interpretation",
            "",
            "The external email test set remains separate from model development. Numerical",
            "targets are descriptive rather than a guarantee of representativeness. See the data",
            "card for provenance, privacy, and distribution-shift limitations.",
            "",
        ]
    )
    markdown_path.write_text("\n".join(lines), encoding="utf-8")


def require_dict(value: object, name: str) -> dict[str, object]:
    """Narrow JSON-like report values with a useful runtime error."""

    if not isinstance(value, dict):
        raise TypeError(f"Expected {name} to be an object")
    return cast(dict[str, object], value)


def prepare_all(
    manifest_path: Path,
    raw_dir: Path,
    output_dir: Path,
    report_json: Path,
    report_markdown: Path,
    *,
    seed: int = 20260806,
) -> dict[str, object]:
    """Run the complete deterministic preparation workflow."""

    sources = load_sources(manifest_path)
    source_stats = {source.id: SourceStats() for source in sources}
    raw_records: list[RawRecord] = []
    for source in sources:
        path = raw_dir / source.id / source.filename
        if not path.exists():
            raise FileNotFoundError(f"Missing {path}; run scripts/download_data.py first")
        verify_file(path, source.checksum, source.expected_bytes)
        raw_records.extend(load_records(source, path, source_stats[source.id]))

    unique_records, duplicates, label_conflicts = deduplicate(raw_records)
    for record in unique_records:
        source_stats[record.source].retained_rows += 1
    processed_records = process_records(unique_records, seed)
    output_manifest = write_outputs(processed_records, output_dir, seed)
    report = build_quality_report(
        sources, source_stats, output_manifest, duplicates, label_conflicts
    )
    write_reports(report, report_json, report_markdown)
    return report


def main() -> None:
    """Run dataset preparation from the command line."""

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--manifest", type=Path, default=Path("data/sources.json"))
    parser.add_argument("--raw-dir", type=Path, default=Path("data/raw"))
    parser.add_argument("--output-dir", type=Path, default=Path("data/processed"))
    parser.add_argument("--report-json", type=Path, default=Path("reports/data-quality.json"))
    parser.add_argument("--report-markdown", type=Path, default=Path("docs/data-quality-report.md"))
    parser.add_argument("--seed", type=int, default=20260806)
    args = parser.parse_args()
    report = prepare_all(
        args.manifest,
        args.raw_dir,
        args.output_dir,
        args.report_json,
        args.report_markdown,
        seed=args.seed,
    )
    processed = require_dict(report["processed"], "processed")
    print(f"prepared {processed['total_records']} records")


if __name__ == "__main__":
    main()
